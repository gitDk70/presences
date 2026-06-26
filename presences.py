#!/usr/bin/env python3
"""Attendance tracking app — generates a styled Excel report."""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from collections import OrderedDict

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _resource_path(filename: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, filename)


class AttendanceApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Présences")
        self.root.geometry("700x720")
        self.root.resizable(True, True)

        # Date frozen at startup
        self.event_date = datetime.now().strftime("%d/%m/%Y")

        # Main records: appt -> {debut, fin, comment}
        self.records: OrderedDict[str, dict] = OrderedDict()

        # Pending queues: list of (appt, time, comment)
        self.debut_pending: list[tuple[str, str, str]] = []
        self.fin_pending:   list[tuple[str, str, str]] = []

        # Time-capture guards
        self._debut_time_set = False
        self._fin_time_set   = False

        self._build_ui()

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        # Logo header
        try:
            self._logo_img = tk.PhotoImage(file=_resource_path("user.png"))
            logo_lbl = ttk.Label(self.root, image=self._logo_img)
            logo_lbl.pack(pady=(10, 2))
        except Exception:
            pass

        # Event header
        info = ttk.LabelFrame(self.root, text="Informations", padding=10)
        info.pack(fill="x", padx=12, pady=(10, 4))
        info.columnconfigure(1, weight=1)

        ttk.Label(info, text="Évènement :").grid(row=0, column=0, sticky="w")
        self.event_var = tk.StringVar()
        ttk.Entry(info, textvariable=self.event_var, width=34).grid(
            row=0, column=1, sticky="ew", padx=(8, 20)
        )
        ttk.Label(info, text="Date :").grid(row=0, column=2, sticky="e")
        ttk.Label(
            info,
            text=self.event_date,
            relief="sunken",
            width=12,
            anchor="center",
            padding=(4, 2),
        ).grid(row=0, column=3, padx=(6, 0))

        # Tabs
        nb = ttk.Notebook(self.root)
        nb.pack(fill="x", padx=12, pady=6)

        tab1 = ttk.Frame(nb, padding=14)
        tab2 = ttk.Frame(nb, padding=14)
        nb.add(tab1, text="Début")
        nb.add(tab2, text="Fin")

        self._build_tab(
            tab1,
            appt_var_name="debut_appt_var",
            time_var_name="debut_time_var",
            comment_var_name="debut_comment_var",
            time_col="Début",
            on_change=self._on_debut_change,
            on_add=self._add_to_debut,
            on_submit=self._record_debut,
            pending_tree_attr="debut_tree",
        )
        self._build_tab(
            tab2,
            appt_var_name="fin_appt_var",
            time_var_name="fin_time_var",
            comment_var_name="fin_comment_var",
            time_col="Fin",
            on_change=self._on_fin_change,
            on_add=self._add_to_fin,
            on_submit=self._record_fin,
            pending_tree_attr="fin_tree",
        )

        # Action bar
        action_bar = ttk.Frame(self.root)
        action_bar.pack(fill="x", padx=12, pady=(4, 2))

        self.count_label = ttk.Label(
            action_bar, text="Nombre de présents : 0", font=("TkDefaultFont", 10, "bold")
        )
        self.count_label.pack(side="left")

        ttk.Button(
            action_bar,
            text="Générer",
            command=self._generate_excel,
            padding=(18, 6),
        ).pack(side="right")

        # Search bar
        search_frame = ttk.Frame(self.root)
        search_frame.pack(fill="x", padx=12, pady=(2, 0))

        ttk.Label(search_frame, text="Rechercher :").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._refresh_table())
        ttk.Entry(search_frame, textvariable=self.search_var, width=28).pack(
            side="left", padx=6
        )
        ttk.Button(
            search_frame,
            text="Effacer",
            command=lambda: self.search_var.set(""),
        ).pack(side="left")

        # Main records table
        tbl = ttk.LabelFrame(self.root, text="Présences enregistrées", padding=6)
        tbl.pack(fill="both", expand=True, padx=12, pady=(4, 10))

        cols = ("Appartement", "Début", "Fin", "Commentaire")
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings", height=7)
        self.tree.heading("Appartement", text="Appartement")
        self.tree.heading("Début",       text="Début")
        self.tree.heading("Fin",         text="Fin")
        self.tree.heading("Commentaire", text="Commentaire")
        self.tree.column("Appartement", width=140, anchor="center")
        self.tree.column("Début",       width=80,  anchor="center")
        self.tree.column("Fin",         width=80,  anchor="center")
        self.tree.column("Commentaire", width=280, anchor="w")

        sb = ttk.Scrollbar(tbl, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Copyright footer
        ttk.Label(
            self.root,
            text="© 2026 Djaouida Kharchi",
            foreground="gray",
            font=("TkDefaultFont", 9),
        ).pack(pady=(0, 6))

    def _build_tab(
        self,
        parent: ttk.Frame,
        *,
        appt_var_name: str,
        time_var_name: str,
        comment_var_name: str,
        time_col: str,
        on_change,
        on_add,
        on_submit,
        pending_tree_attr: str,
    ) -> None:
        parent.columnconfigure(1, weight=1)

        # Row 0: Appartement + time
        ttk.Label(parent, text="Appartement :").grid(row=0, column=0, sticky="w", pady=3)
        appt_var = tk.StringVar()
        appt_var.trace_add("write", on_change)
        setattr(self, appt_var_name, appt_var)

        entry = ttk.Entry(parent, textvariable=appt_var, width=24)
        entry.grid(row=0, column=1, padx=(8, 20), pady=3, sticky="w")
        entry.bind("<Return>", on_add)

        ttk.Label(parent, text=f"{time_col} :").grid(row=0, column=2, sticky="w")
        time_var = tk.StringVar()
        setattr(self, time_var_name, time_var)
        ttk.Entry(parent, textvariable=time_var, width=10, state="readonly").grid(
            row=0, column=3, padx=(4, 0), pady=3, sticky="w"
        )

        # Row 1: Comment
        ttk.Label(parent, text="Commentaire :").grid(row=1, column=0, sticky="w", pady=3)
        comment_var = tk.StringVar()
        setattr(self, comment_var_name, comment_var)
        comment_entry = ttk.Entry(parent, textvariable=comment_var, width=48)
        comment_entry.grid(row=1, column=1, columnspan=3, padx=(8, 0), pady=3, sticky="ew")
        comment_entry.bind("<Return>", on_add)

        # Row 2: Hint
        ttk.Label(
            parent,
            text="Entrée pour ajouter · Suppr pour retirer",
            foreground="gray",
        ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(2, 4))

        # Row 3: Pending treeview
        pending_tree = ttk.Treeview(
            parent,
            columns=("Appartement", time_col, "Commentaire"),
            show="headings",
            height=4,
        )
        pending_tree.heading("Appartement", text="Appartement")
        pending_tree.heading(time_col,      text=time_col)
        pending_tree.heading("Commentaire", text="Commentaire")
        pending_tree.column("Appartement", width=140, anchor="center")
        pending_tree.column(time_col,      width=80,  anchor="center")
        pending_tree.column("Commentaire", width=280, anchor="w")
        pending_tree.grid(row=3, column=0, columnspan=4, sticky="ew", pady=4)
        setattr(self, pending_tree_attr, pending_tree)

        pending_list = self.debut_pending if "debut" in appt_var_name else self.fin_pending
        for seq in ("<Delete>", "<BackSpace>"):
            pending_tree.bind(
                seq,
                lambda e, t=pending_tree, p=pending_list: self._delete_selected(t, p),
            )

        # Row 4: Buttons
        btn_frame = ttk.Frame(parent)
        btn_frame.grid(row=4, column=0, columnspan=4, sticky="e", pady=(4, 0))

        ttk.Button(
            btn_frame,
            text="Supprimer la sélection",
            command=lambda t=pending_tree, p=pending_list: self._delete_selected(t, p),
        ).pack(side="left", padx=(0, 8))

        ttk.Button(btn_frame, text="Enregistrer", command=on_submit).pack(side="left")

    # ── Auto-fill time on first character ─────────────────────────────────────

    def _on_debut_change(self, *_) -> None:
        val = self.debut_appt_var.get()
        if val and not self._debut_time_set:
            self.debut_time_var.set(datetime.now().strftime("%H:%M"))
            self._debut_time_set = True
        elif not val:
            self._debut_time_set = False
            self.debut_time_var.set("")

    def _on_fin_change(self, *_) -> None:
        val = self.fin_appt_var.get()
        if val and not self._fin_time_set:
            self.fin_time_var.set(datetime.now().strftime("%H:%M"))
            self._fin_time_set = True
        elif not val:
            self._fin_time_set = False
            self.fin_time_var.set("")

    # ── Add to pending queue (Enter key) ──────────────────────────────────────

    def _add_to_debut(self, *_) -> None:
        appt = self.debut_appt_var.get().strip()
        if not appt:
            return
        t       = self.debut_time_var.get() or datetime.now().strftime("%H:%M")
        comment = self.debut_comment_var.get().strip()
        self.debut_pending.append((appt, t, comment))
        self.debut_tree.insert("", "end", values=(appt, t, comment))
        self.debut_appt_var.set("")
        self.debut_time_var.set("")
        self.debut_comment_var.set("")
        self._debut_time_set = False

    def _add_to_fin(self, *_) -> None:
        appt = self.fin_appt_var.get().strip()
        if not appt:
            return
        t       = self.fin_time_var.get() or datetime.now().strftime("%H:%M")
        comment = self.fin_comment_var.get().strip()
        self.fin_pending.append((appt, t, comment))
        self.fin_tree.insert("", "end", values=(appt, t, comment))
        self.fin_appt_var.set("")
        self.fin_time_var.set("")
        self.fin_comment_var.set("")
        self._fin_time_set = False

    # ── Delete selected row from pending treeview ─────────────────────────────

    def _delete_selected(self, tree: ttk.Treeview, pending: list) -> None:
        for item in tree.selection():
            idx = tree.index(item)
            tree.delete(item)
            if 0 <= idx < len(pending):
                pending.pop(idx)

    # ── Enregistrer — flush pending queue to main records ─────────────────────

    def _record_debut(self, *_) -> None:
        self._add_to_debut()
        if not self.debut_pending:
            messagebox.showwarning("Attention", "Aucun appartement à enregistrer.")
            return
        for appt, t, comment in self.debut_pending:
            if appt in self.records:
                self.records[appt]["debut"]   = t
                self.records[appt]["comment"] = comment
            else:
                self.records[appt] = {"debut": t, "fin": "", "comment": comment}
        self.debut_pending.clear()
        self.debut_tree.delete(*self.debut_tree.get_children())
        self._refresh_table()

    def _record_fin(self, *_) -> None:
        self._add_to_fin()
        if not self.fin_pending:
            messagebox.showwarning("Attention", "Aucun appartement à enregistrer.")
            return
        for appt, t, comment in self.fin_pending:
            if appt in self.records:
                self.records[appt]["fin"] = t
                if comment:
                    self.records[appt]["comment"] = comment
            else:
                self.records[appt] = {"debut": "", "fin": t, "comment": comment}
        self.fin_pending.clear()
        self.fin_tree.delete(*self.fin_tree.get_children())
        self._refresh_table()

    # ── Refresh main table (applies search filter) ────────────────────────────

    def _refresh_table(self) -> None:
        self.tree.delete(*self.tree.get_children())
        query = self.search_var.get().strip().lower()
        for appt, times in self.records.items():
            if query and query not in appt.lower():
                continue
            self.tree.insert(
                "", "end",
                values=(appt, times["debut"], times["fin"], times.get("comment", "")),
            )
        self.count_label.config(text=f"Nombre de présents : {len(self.records)}")

    # ── Excel export ──────────────────────────────────────────────────────────

    def _generate_excel(self) -> None:
        if not self.records:
            messagebox.showwarning("Attention", "Aucune présence enregistrée.")
            return

        event_name = self.event_var.get().strip()
        safe    = event_name.replace(" ", "_") or "evenement"
        default = f"presences_{safe}_{self.event_date.replace('/', '-')}.xlsx"

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")],
            initialfile=default,
            title="Enregistrer le fichier Excel",
        )
        self.root.lift()
        self.root.focus_force()
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Présences"

        bold     = Font(bold=True)
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", fgColor="2E5FA3")
        center   = Alignment(horizontal="center", vertical="center")
        left     = Alignment(horizontal="left",   vertical="center")
        thin     = Side(style="thin")
        border   = Border(left=thin, right=thin, top=thin, bottom=thin)

        def cell(row, col, value, *, font=None, fill=None, align=None, brd=None):
            c = ws.cell(row=row, column=col, value=value)
            if font:  c.font      = font
            if fill:  c.fill      = fill
            if align: c.alignment = align
            if brd:   c.border    = brd
            return c

        cell(1, 1, "Évènement :", font=bold)
        cell(1, 2, event_name)
        cell(1, 4, "Date :", font=bold)
        cell(1, 5, self.event_date)

        cell(3, 1, "Nombre de présents :", font=bold)
        cell(3, 2, len(self.records))

        for col, header in enumerate(["Appartement", "Début", "Fin", "Commentaire"], start=1):
            cell(5, col, header, font=hdr_font, fill=hdr_fill, align=center, brd=border)

        for row_idx, (appt, times) in enumerate(self.records.items(), start=6):
            cell(row_idx, 1, appt,                       align=center, brd=border)
            cell(row_idx, 2, times["debut"],              align=center, brd=border)
            cell(row_idx, 3, times["fin"],                align=center, brd=border)
            cell(row_idx, 4, times.get("comment", ""),   align=left,   brd=border)

        for col, width in zip("ABCDE", [22, 12, 12, 35, 14]):
            ws.column_dimensions[col].width = width

        wb.save(path)
        messagebox.showinfo("Succès", f"Fichier généré :\n{path}")
        self.root.lift()
        self.root.focus_force()


def main() -> None:
    root = tk.Tk()
    AttendanceApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
